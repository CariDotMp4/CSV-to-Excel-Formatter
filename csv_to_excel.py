"""
csv_to_excel.py
---------------
1. Preencha os campos e carregue o CSV  →  "Carregar e Editar"
2. Edite a tabela: remova colunas (clique no cabeçalho) ou linhas (clique no nº da linha)
3. Clique em "Gerar Planilha"

Gerar executável:
    python -m PyInstaller --onefile --windowed --name "CSV_Excel_Formatter" csv_to_excel.py
"""

import argparse
import os
import platform
import subprocess
import sys
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
#  ESTILOS DA PLANILHA (openpyxl)
# ══════════════════════════════════════════════════════════════════════════════

class Estilos:
    ROXO        = "7030A0"
    ROXO_CLARO  = "E2D0F0"
    BRANCO      = "FFFFFF"
    PRETO       = "000000"
    BORDA       = "7030A0"
    FONTE       = "Arial"
    T_TITULO    = 16
    T_SUB       = 10
    T_CAB       = 11
    T_DADOS     = 10
    T_RODAPE    = 10
    L_TITULO    = 1
    L_SUB       = 2
    L_BRANCO    = 3
    L_CAB       = 4


def _borda(cor=Estilos.BORDA):
    s = Side(style="thin", color=cor)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(cor):
    return PatternFill("solid", fgColor=cor)

def _font(cor, size, bold=False):
    return Font(name=Estilos.FONTE, size=size, bold=bold, color=cor)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ══════════════════════════════════════════════════════════════════════════════
#  LEITURA DO CSV
# ══════════════════════════════════════════════════════════════════════════════

def ler_csv(caminho: Path, sep=",", enc="utf-8") -> pd.DataFrame:
    try:
        df = pd.read_csv(caminho, sep=sep, encoding=enc, skipinitialspace=True)
    except UnicodeDecodeError:
        df = pd.read_csv(caminho, sep=sep, encoding="latin-1", skipinitialspace=True)
    df.dropna(how="all", axis=1, inplace=True)
    df.dropna(how="all", axis=0, inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  GERAÇÃO DO EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def _col_width(df, col):
    lens = [len(str(col))] + [len(str(v)) for v in df[col] if pd.notna(v)]
    return min(max(max(lens) + 2, 10), 60)


def gerar_excel(df: pd.DataFrame, caminho_saida: str, titulo: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    n_cols  = len(df.columns)
    n_rows  = len(df)
    col_fim = get_column_letter(n_cols)

    # Título
    ws.merge_cells(f"A{Estilos.L_TITULO}:{col_fim}{Estilos.L_TITULO}")
    ct = ws[f"A{Estilos.L_TITULO}"]
    ct.value = titulo
    ct.font  = _font(Estilos.ROXO, Estilos.T_TITULO, bold=True)
    ct.alignment = _align("center")
    ws.row_dimensions[Estilos.L_TITULO].height = 34

    # Subtítulo
    ws.merge_cells(f"A{Estilos.L_SUB}:{col_fim}{Estilos.L_SUB}")
    cs = ws[f"A{Estilos.L_SUB}"]
    cs.value = f"{n_rows} registro(s)  ·  {n_cols} coluna(s)"
    cs.font  = _font(Estilos.ROXO, Estilos.T_SUB)
    cs.alignment = _align("center")
    ws.row_dimensions[Estilos.L_SUB].height = 18
    ws.row_dimensions[Estilos.L_BRANCO].height = 8

    # Cabeçalho
    for c, nome in enumerate(df.columns, 1):
        cel = ws.cell(row=Estilos.L_CAB, column=c, value=str(nome).strip())
        cel.font      = _font(Estilos.BRANCO, Estilos.T_CAB, bold=True)
        cel.fill      = _fill(Estilos.ROXO)
        cel.alignment = _align("center", wrap=True)
        cel.border    = _borda()
    ws.row_dimensions[Estilos.L_CAB].height = 30

    # Dados
    l0 = Estilos.L_CAB + 1
    for r, row in enumerate(df.itertuples(index=False)):
        linha = l0 + r
        fundo = Estilos.ROXO_CLARO if r % 2 == 0 else Estilos.BRANCO
        for c, val in enumerate(row, 1):
            v   = "" if pd.isna(val) else val
            cel = ws.cell(row=linha, column=c, value=v)
            cel.font      = _font(Estilos.PRETO, Estilos.T_DADOS)
            cel.fill      = _fill(fundo)
            cel.border    = _borda()
            cel.alignment = (
                _align("right") if isinstance(val, (int, float)) and pd.notna(val)
                else _align("left")
            )
        ws.row_dimensions[linha].height = 18

    # Rodapé
    lr = l0 + n_rows
    ws.row_dimensions[lr].height = 22
    for c, nome in enumerate(df.columns, 1):
        cel = ws.cell(row=lr, column=c)
        cel.fill      = _fill(Estilos.ROXO)
        cel.font      = _font(Estilos.BRANCO, Estilos.T_RODAPE, bold=True)
        cel.border    = _borda()
        cel.alignment = _align("center")
        if pd.api.types.is_numeric_dtype(df[nome]):
            cl = get_column_letter(c)
            cel.value = f"=SUM({cl}{l0}:{cl}{lr-1})"
        elif c == 1:
            cel.value = f"Total: {n_rows} registros"

    # Larguras
    for c, nome in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(c)].width = _col_width(df, nome)

    ws.freeze_panes = f"A{l0}"
    ws.auto_filter.ref = f"A{Estilos.L_CAB}:{col_fim}{lr-1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"{Estilos.L_CAB}:{Estilos.L_CAB}"

    # Aba Info
    wi = wb.create_sheet("Info")
    for i, (k, v) in enumerate([
        ("Gerado por", "csv_to_excel"),
        ("Registros",  n_rows),
        ("Colunas",    n_cols),
        ("Lista de colunas", ", ".join(df.columns.tolist())),
    ], 1):
        wi.cell(i, 1, k).font = Font(bold=True)
        wi.cell(i, 2, v)
    wi.column_dimensions["A"].width = 20
    wi.column_dimensions["B"].width = 70

    wb.save(caminho_saida)
    return caminho_saida


# ══════════════════════════════════════════════════════════════════════════════
#  CORES DA UI
# ══════════════════════════════════════════════════════════════════════════════

ROXO_UI      = "#7030A0"
ROXO_HOVER   = "#5a2080"
ROXO_CLARO   = "#f3e8fc"
ROXO_BG      = "#faf5ff"
TEXTO        = "#1a1a2e"
CINZA_BORDA  = "#d8b4fe"
BRANCO       = "#ffffff"
VERDE_OK     = "#16a34a"
VERMELHO_ERR = "#dc2626"
AMARELO_SEL  = "#fef9c3"   # highlight de linha/col marcada para remoção

PH_CSV    = "Caminho completo do arquivo .csv"
PH_SAIDA  = "Caminho onde o Excel será salvo"
PH_TITULO = "Ex: Relatório de Vendas 2025"


# ══════════════════════════════════════════════════════════════════════════════
#  WIDGET: EDITOR DE TABELA
# ══════════════════════════════════════════════════════════════════════════════

class TabelaEditor(tk.Frame):
    """
    Grade editável que exibe o DataFrame.
    • Clique no cabeçalho de coluna  → marca/desmarca a coluna para remoção (fundo vermelho)
    • Clique no número de linha      → marca/desmarca a linha para remoção (fundo vermelho)
    • Botão "Remover marcados"       → aplica as remoções ao DataFrame interno
    A coluna de índice (nº da linha) não é editável.
    """

    COR_CAB_NORMAL  = ROXO_UI
    COR_CAB_MARCADO = "#c0392b"
    COR_LIN_PAR     = "#f5f0ff"
    COR_LIN_IMPAR   = BRANCO
    COR_LIN_MARCADA = "#fde8e8"
    COR_IDX         = "#ede9f6"   # coluna de índice

    def __init__(self, master, **kw):
        super().__init__(master, bg=ROXO_BG, **kw)
        self.df: pd.DataFrame | None = None
        self._cols_marcadas: set[str] = set()
        self._linhas_marcadas: set[int] = set()   # índices do df
        self._widgets_cab  = {}   # col_name → Button
        self._widgets_idx  = {}   # df_idx   → Button
        self._widgets_cel  = {}   # (df_idx, col_name) → Entry

        self._build_vazio()

    # ── estrutura base ────────────────────────────────────────────────────────

    def _build_vazio(self):
        self._info = tk.Label(self, text="Nenhum arquivo carregado.",
                              bg=ROXO_BG, fg="#9ca3af",
                              font=("Arial", 10, "italic"))
        self._info.pack(expand=True)

    def _limpar(self):
        for w in self.winfo_children():
            w.destroy()
        self._widgets_cab.clear()
        self._widgets_idx.clear()
        self._widgets_cel.clear()

    # ── carregamento ──────────────────────────────────────────────────────────

    def carregar(self, df: pd.DataFrame):
        self.df = df.copy()
        self._cols_marcadas.clear()
        self._linhas_marcadas.clear()
        self._renderizar()

    def _renderizar(self):
        self._limpar()

        if self.df is None or self.df.empty:
            self._build_vazio()
            return

        # ── barra de ações ────────────────────────────────────────────────────
        barra = tk.Frame(self, bg=ROXO_BG)
        barra.pack(fill="x", pady=(0, 4))

        self._lbl_sel = tk.Label(barra,
                                 text=self._texto_selecao(),
                                 bg=ROXO_BG, fg=TEXTO,
                                 font=("Arial", 9))
        self._lbl_sel.pack(side="left")

        tk.Button(barra, text="↩  Restaurar tudo",
                  bg=ROXO_CLARO, fg=ROXO_UI,
                  font=("Arial", 9, "bold"), relief="flat", cursor="hand2",
                  activebackground=CINZA_BORDA,
                  command=self._restaurar
                  ).pack(side="right", padx=(6, 0))

        tk.Button(barra, text="🗑  Remover marcados",
                  bg="#fde8e8", fg=self.COR_CAB_MARCADO,
                  font=("Arial", 9, "bold"), relief="flat", cursor="hand2",
                  activebackground="#fca5a5",
                  command=self._aplicar_remocoes
                  ).pack(side="right")

        # ── área com scroll ───────────────────────────────────────────────────
        cont = tk.Frame(self, bg=ROXO_BG)
        cont.pack(fill="both", expand=True)

        vbar = ttk.Scrollbar(cont, orient="vertical")
        hbar = ttk.Scrollbar(cont, orient="horizontal")
        vbar.pack(side="right",  fill="y")
        hbar.pack(side="bottom", fill="x")

        canvas = tk.Canvas(cont, bg=ROXO_BG,
                           yscrollcommand=vbar.set,
                           xscrollcommand=hbar.set,
                           highlightthickness=0)
        canvas.pack(side="left", fill="both", expand=True)

        vbar.config(command=canvas.yview)
        hbar.config(command=canvas.xview)

        grade = tk.Frame(canvas, bg=ROXO_BG)
        canvas_win = canvas.create_window((0, 0), window=grade, anchor="nw")

        def _on_resize(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        grade.bind("<Configure>", _on_resize)

        # Scroll com mouse
        def _scroll_y(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _scroll_y)

        # ── cabeçalhos ────────────────────────────────────────────────────────
        # célula vazia no canto (linha 0, col 0)
        tk.Label(grade, text="#", bg=self.COR_IDX, fg=TEXTO,
                 font=("Arial", 9, "bold"),
                 width=4, relief="flat",
                 padx=4, pady=4
                 ).grid(row=0, column=0, sticky="nsew", padx=1, pady=1)

        for c_idx, col in enumerate(self.df.columns, 1):
            marcada = col in self._cols_marcadas
            bg = self.COR_CAB_MARCADO if marcada else self.COR_CAB_NORMAL
            txt = f"✕ {col}" if marcada else col
            btn = tk.Button(grade, text=txt,
                            bg=bg, fg=BRANCO,
                            font=("Arial", 9, "bold"),
                            relief="flat", cursor="hand2",
                            wraplength=110, justify="center",
                            padx=6, pady=4,
                            command=lambda c=col: self._toggle_col(c))
            btn.grid(row=0, column=c_idx, sticky="nsew", padx=1, pady=1)
            self._widgets_cab[col] = btn

        # ── linhas de dados ───────────────────────────────────────────────────
        for r_idx, (df_idx, row) in enumerate(self.df.iterrows(), 1):
            marcada_lin = df_idx in self._linhas_marcadas
            bg_lin = self.COR_LIN_MARCADA if marcada_lin else (
                self.COR_LIN_PAR if r_idx % 2 == 0 else self.COR_LIN_IMPAR
            )

            # Botão de índice
            txt_idx = f"✕ {r_idx}" if marcada_lin else str(r_idx)
            btn_idx = tk.Button(grade, text=txt_idx,
                                bg=self.COR_CAB_MARCADO if marcada_lin else self.COR_IDX,
                                fg=BRANCO if marcada_lin else TEXTO,
                                font=("Arial", 9),
                                relief="flat", cursor="hand2",
                                width=4, padx=2, pady=2,
                                command=lambda i=df_idx: self._toggle_linha(i))
            btn_idx.grid(row=r_idx, column=0, sticky="nsew", padx=1, pady=1)
            self._widgets_idx[df_idx] = btn_idx

            for c_idx, col in enumerate(self.df.columns, 1):
                val = row[col]
                txt = "" if pd.isna(val) else str(val)
                var = tk.StringVar(value=txt)

                entry = tk.Entry(grade, textvariable=var,
                                 bg=bg_lin, fg=TEXTO,
                                 font=("Arial", 9),
                                 relief="flat",
                                 highlightthickness=1,
                                 highlightbackground=CINZA_BORDA,
                                 highlightcolor=ROXO_UI,
                                 width=14)
                entry.grid(row=r_idx, column=c_idx, sticky="nsew", padx=1, pady=1)
                # Salva edições de volta no df ao sair da célula
                entry.bind("<FocusOut>",
                           lambda e, i=df_idx, c=col, v=var: self._salvar_cell(i, c, v))
                entry.bind("<Return>",
                           lambda e, i=df_idx, c=col, v=var: self._salvar_cell(i, c, v))
                self._widgets_cel[(df_idx, col)] = (entry, var)

    # ── ações de seleção ─────────────────────────────────────────────────────

    def _toggle_col(self, col: str):
        if col in self._cols_marcadas:
            self._cols_marcadas.discard(col)
        else:
            self._cols_marcadas.add(col)
        self._renderizar()

    def _toggle_linha(self, df_idx: int):
        if df_idx in self._linhas_marcadas:
            self._linhas_marcadas.discard(df_idx)
        else:
            self._linhas_marcadas.add(df_idx)
        self._renderizar()

    def _restaurar(self):
        self._cols_marcadas.clear()
        self._linhas_marcadas.clear()
        self._renderizar()

    def _aplicar_remocoes(self):
        if not self._cols_marcadas and not self._linhas_marcadas:
            return
        n_col = len(self._cols_marcadas)
        n_lin = len(self._linhas_marcadas)
        msg = []
        if n_col: msg.append(f"{n_col} coluna(s)")
        if n_lin: msg.append(f"{n_lin} linha(s)")
        if not messagebox.askyesno("Confirmar remoção",
                                   f"Remover permanentemente {' e '.join(msg)}?\n\nEsta ação não pode ser desfeita."):
            return
        if self._linhas_marcadas:
            self.df = self.df.drop(index=list(self._linhas_marcadas)).reset_index(drop=True)
        if self._cols_marcadas:
            self.df = self.df.drop(columns=list(self._cols_marcadas))
        self._cols_marcadas.clear()
        self._linhas_marcadas.clear()
        self._renderizar()

    def _salvar_cell(self, df_idx, col, var: tk.StringVar):
        """Atualiza o valor no DataFrame quando o usuário edita uma célula."""
        novo = var.get()
        try:
            # Tenta preservar o tipo numérico
            if pd.api.types.is_numeric_dtype(self.df[col]):
                self.df.at[df_idx, col] = float(novo) if novo else None
            else:
                self.df.at[df_idx, col] = novo
        except Exception:
            self.df.at[df_idx, col] = novo

    def _texto_selecao(self):
        partes = []
        if self._cols_marcadas:
            partes.append(f"{len(self._cols_marcadas)} coluna(s) marcada(s)")
        if self._linhas_marcadas:
            partes.append(f"{len(self._linhas_marcadas)} linha(s) marcada(s)")
        return "  ·  ".join(partes) if partes else "Clique em cabeçalhos ou números de linha para marcar para remoção."

    def df_editado(self) -> pd.DataFrame | None:
        return self.df


# ══════════════════════════════════════════════════════════════════════════════
#  JANELA PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CSV → Excel Formatter")
        self.configure(bg=ROXO_BG)
        self.minsize(900, 640)
        self._center(1060, 720)
        self._build()

    def _center(self, w, h):
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = max(0, (sw - w) // 2)
        y = max(0, (sh - h) // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")

    # ── construção ────────────────────────────────────────────────────────────

    def _build(self):
        # Cabeçalho
        header = tk.Frame(self, bg=ROXO_UI, height=64)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="CSV → Excel Formatter",
                 bg=ROXO_UI, fg=BRANCO,
                 font=("Arial", 17, "bold")).pack(side="left", padx=24, pady=14)
        tk.Label(header, text="SharePoint / Local",
                 bg=ROXO_UI, fg="#d8b4fe",
                 font=("Arial", 10)).pack(side="right", padx=24)

        # Painel superior: campos de config
        painel = tk.Frame(self, bg=ROXO_BG, padx=24, pady=14)
        painel.pack(fill="x")

        # Linha 1: CSV + Saída + Separador
        l1 = tk.Frame(painel, bg=ROXO_BG)
        l1.pack(fill="x")

        self.csv_path   = tk.StringVar()
        self.saida_path = tk.StringVar()
        self.titulo_var = tk.StringVar()
        self.sep_var    = tk.StringVar(value=",")

        self._campo_inline(l1, "📄 CSV de entrada",  self.csv_path,  PH_CSV,
                           btn="Procurar…", cmd=self._browse_csv,   flex=True)
        tk.Frame(l1, bg=ROXO_BG, width=12).pack(side="left")
        self._campo_inline(l1, "💾 Saída (.xlsx)",   self.saida_path, PH_SAIDA,
                           btn="Salvar em…", cmd=self._browse_saida, flex=True)
        tk.Frame(l1, bg=ROXO_BG, width=12).pack(side="left")
        self._campo_inline(l1, "🔤 Separador", self.sep_var, "",
                           width=5, flex=False)

        # Linha 2: Título + botão Carregar
        l2 = tk.Frame(painel, bg=ROXO_BG)
        l2.pack(fill="x", pady=(10, 0))

        self._campo_inline(l2, "🏷️ Título da planilha", self.titulo_var, PH_TITULO,
                           flex=True)

        tk.Frame(l2, bg=ROXO_BG, width=12).pack(side="left")
        self.btn_carregar = tk.Button(l2, text="⬆  Carregar e Editar",
                                      bg=ROXO_UI, fg=BRANCO,
                                      font=("Arial", 10, "bold"),
                                      relief="flat", cursor="hand2",
                                      activebackground=ROXO_HOVER,
                                      padx=16, pady=7,
                                      command=self._carregar)
        self.btn_carregar.pack(side="left", anchor="s")
        self.btn_carregar.bind("<Enter>", lambda e: self.btn_carregar.config(bg=ROXO_HOVER))
        self.btn_carregar.bind("<Leave>", lambda e: self.btn_carregar.config(bg=ROXO_UI))

        # Divisor
        tk.Frame(self, bg=CINZA_BORDA, height=1).pack(fill="x")

        # Editor de tabela (ocupa o espaço restante)
        self.editor = TabelaEditor(self)
        self.editor.pack(fill="both", expand=True, padx=16, pady=(10, 0))

        # Rodapé: status + progresso + botão gerar
        rodape = tk.Frame(self, bg=ROXO_BG, padx=24, pady=10)
        rodape.pack(fill="x", side="bottom")

        self.status_var = tk.StringVar(value="Carregue um CSV para começar.")
        self.status_lbl = tk.Label(rodape, textvariable=self.status_var,
                                   bg=ROXO_BG, fg=TEXTO,
                                   font=("Arial", 9), anchor="w")
        self.status_lbl.pack(fill="x")

        style = ttk.Style(self)
        style.theme_use("default")
        style.configure("P.Horizontal.TProgressbar",
                        troughcolor=ROXO_CLARO, background=ROXO_UI, thickness=5)
        self.progress = ttk.Progressbar(rodape, mode="indeterminate",
                                        style="P.Horizontal.TProgressbar")
        self.progress.pack(fill="x", pady=(4, 8))

        self.btn_gerar = tk.Button(rodape, text="✨  Gerar Planilha",
                                   bg=ROXO_UI, fg=BRANCO,
                                   font=("Arial", 12, "bold"),
                                   relief="flat", cursor="hand2",
                                   activebackground=ROXO_HOVER,
                                   padx=28, pady=10,
                                   state="disabled",
                                   command=self._gerar)
        self.btn_gerar.pack()
        self.btn_gerar.bind("<Enter>", lambda e: self.btn_gerar.config(bg=ROXO_HOVER) if self.btn_gerar["state"] == "normal" else None)
        self.btn_gerar.bind("<Leave>", lambda e: self.btn_gerar.config(bg=ROXO_UI))

    def _campo_inline(self, parent, label, var, placeholder,
                      btn=None, cmd=None, width=None, flex=False):
        """Cria label + entry (+ botão) horizontalmente num frame pai."""
        frm = tk.Frame(parent, bg=ROXO_BG)
        frm.pack(side="left", fill="x", expand=flex)

        tk.Label(frm, text=label, bg=ROXO_BG, fg=TEXTO,
                 font=("Arial", 8, "bold")).pack(anchor="w")

        row = tk.Frame(frm, bg=ROXO_BG)
        row.pack(fill="x", pady=(3, 0))

        kw = {"width": width} if width else {}
        entry = tk.Entry(row, textvariable=var,
                         font=("Arial", 10),
                         bg=BRANCO, fg="#9ca3af",
                         relief="flat",
                         highlightthickness=1,
                         highlightbackground=CINZA_BORDA,
                         highlightcolor=ROXO_UI,
                         insertbackground=ROXO_UI, **kw)
        entry.pack(side="left", fill="x", expand=flex, ipady=6)

        if placeholder:
            entry.insert(0, placeholder)
            entry.bind("<FocusIn>",  lambda e, en=entry, ph=placeholder: self._ph_in(en, ph))
            entry.bind("<FocusOut>", lambda e, en=entry, ph=placeholder: self._ph_out(en, ph))

        if btn and cmd:
            tk.Button(row, text=btn,
                      bg=ROXO_CLARO, fg=ROXO_UI,
                      font=("Arial", 9, "bold"), relief="flat", cursor="hand2",
                      activebackground=CINZA_BORDA,
                      padx=8,
                      command=cmd
                      ).pack(side="left", padx=(6, 0), ipady=6)

    # ── placeholders ─────────────────────────────────────────────────────────

    def _ph_in(self, entry, ph):
        if entry.get() == ph:
            entry.delete(0, "end")
            entry.config(fg=TEXTO)

    def _ph_out(self, entry, ph):
        if not entry.get():
            entry.insert(0, ph)
            entry.config(fg="#9ca3af")

    # ── diálogos ─────────────────────────────────────────────────────────────

    def _browse_csv(self):
        p = filedialog.askopenfilename(parent=self, title="Selecionar CSV",
                                       filetypes=[("CSV", "*.csv"), ("Todos", "*.*")])
        if not p:
            return
        self.csv_path.set(p)
        if self.saida_path.get() in ("", PH_SAIDA):
            self.saida_path.set(str(Path(p).with_suffix(".xlsx")))
        if self.titulo_var.get() in ("", PH_TITULO):
            self.titulo_var.set(Path(p).stem.replace("_", " ").title())

    def _browse_saida(self):
        csv = self.csv_path.get()
        inicio = str(Path(csv).parent) if csv and Path(csv).exists() else os.path.expanduser("~")
        p = filedialog.asksaveasfilename(parent=self, title="Salvar como…",
                                          initialdir=inicio,
                                          initialfile="relatorio.xlsx",
                                          defaultextension=".xlsx",
                                          filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")])
        if not p:
            return
        if not p.lower().endswith(".xlsx"):
            p += ".xlsx"
        self.saida_path.set(p)

    # ── status ────────────────────────────────────────────────────────────────

    def _set_status(self, msg, cor=TEXTO):
        self.status_var.set(msg)
        self.status_lbl.config(fg=cor)

    # ── carregar CSV ──────────────────────────────────────────────────────────

    def _carregar(self):
        csv_path = self.csv_path.get().strip()
        sep      = self.sep_var.get().strip() or ","

        if not csv_path or csv_path == PH_CSV:
            self._set_status("⚠️  Informe o caminho do CSV.", VERMELHO_ERR); return
        if not Path(csv_path).exists():
            self._set_status(f"⚠️  Arquivo não encontrado: {csv_path}", VERMELHO_ERR); return

        sep = sep.replace("\\t", "\t")
        self.btn_carregar.config(state="disabled")
        self.btn_gerar.config(state="disabled")
        self.progress.start(12)
        self._set_status("⏳  Lendo CSV…", ROXO_UI)

        def run():
            try:
                df = ler_csv(Path(csv_path), sep=sep)
                self.after(0, self._csv_carregado, df)
            except Exception as exc:
                self.after(0, self._erro_carregar, str(exc))

        threading.Thread(target=run, daemon=True).start()

    def _csv_carregado(self, df: pd.DataFrame):
        self.progress.stop()
        self.btn_carregar.config(state="normal")
        self.btn_gerar.config(state="normal")
        self.editor.carregar(df)
        self._set_status(
            f"✅  CSV carregado: {len(df)} linha(s) × {len(df.columns)} coluna(s). "
            "Clique nos cabeçalhos ou nos números de linha para marcar para remoção.",
            VERDE_OK
        )

    def _erro_carregar(self, msg):
        self.progress.stop()
        self.btn_carregar.config(state="normal")
        self._set_status(f"❌  Erro ao ler CSV: {msg}", VERMELHO_ERR)
        messagebox.showerror("Erro ao carregar", msg)

    # ── gerar Excel ───────────────────────────────────────────────────────────

    def _gerar(self):
        saida  = self.saida_path.get().strip()
        titulo = self.titulo_var.get().strip()

        if not saida or saida == PH_SAIDA:
            self._set_status("⚠️  Informe o caminho de saída.", VERMELHO_ERR); return

        df = self.editor.df_editado()
        if df is None or df.empty:
            self._set_status("⚠️  Nenhum dado para exportar.", VERMELHO_ERR); return

        if not titulo or titulo == PH_TITULO:
            titulo = "Relatório de Dados"

        self.btn_gerar.config(state="disabled")
        self.btn_carregar.config(state="disabled")
        self.progress.start(12)
        self._set_status("⏳  Gerando planilha…", ROXO_UI)

        def run():
            try:
                gerar_excel(df, saida, titulo)
                self.after(0, self._sucesso, saida, len(df), len(df.columns))
            except Exception as exc:
                self.after(0, self._erro_gerar, str(exc))

        threading.Thread(target=run, daemon=True).start()

    def _sucesso(self, saida, linhas, colunas):
        self.progress.stop()
        self.btn_gerar.config(state="normal")
        self.btn_carregar.config(state="normal")
        self._set_status(
            f"✅  Planilha gerada!  {linhas} linha(s) · {colunas} coluna(s)  →  {saida}",
            VERDE_OK
        )
        if messagebox.askyesno("Concluído",
                               f"Planilha gerada com sucesso!\n\n{saida}\n\nDeseja abrir o arquivo?"):
            if platform.system() == "Windows":
                os.startfile(saida)
            elif platform.system() == "Darwin":
                subprocess.call(["open", saida])
            else:
                subprocess.call(["xdg-open", saida])

    def _erro_gerar(self, msg):
        self.progress.stop()
        self.btn_gerar.config(state="normal")
        self.btn_carregar.config(state="normal")
        self._set_status(f"❌  Erro: {msg}", VERMELHO_ERR)
        messagebox.showerror("Erro", msg)


# ══════════════════════════════════════════════════════════════════════════════
#  MODO CLI
# ══════════════════════════════════════════════════════════════════════════════

def cli():
    parser = argparse.ArgumentParser(description="CSV → Excel formatter (terminal)")
    parser.add_argument("--csv",    required=True)
    parser.add_argument("--saida",  required=True)
    parser.add_argument("--titulo", default="Relatório de Dados")
    parser.add_argument("--sep",    default=",")
    parser.add_argument("--enc",    default="utf-8")
    args = parser.parse_args()

    caminho = Path(args.csv)
    if not caminho.exists():
        print(f"❌  Arquivo não encontrado: {caminho}"); sys.exit(1)

    print(f"🔄  Lendo {caminho.name}…")
    df = ler_csv(caminho, sep=args.sep.replace("\\t", "\t"), enc=args.enc)
    print(f"   ✅  {len(df)} linhas × {len(df.columns)} colunas")
    print("📊  Gerando Excel…")
    gerar_excel(df, args.saida, args.titulo)
    print(f"✅  Salvo em: {args.saida}")


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    if "--cli" in sys.argv:
        sys.argv.remove("--cli")
        cli()
    else:
        App().mainloop()
